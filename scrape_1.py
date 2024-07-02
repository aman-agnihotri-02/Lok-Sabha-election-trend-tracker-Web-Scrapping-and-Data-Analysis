from bs4 import BeautifulSoup

import openpyxl.workbook
import requests, openpyxl

excel = openpyxl.Workbook()
print(excel.sheetnames)
sheet = excel.active
sheet.title = 'Parliamentary Constituency General Party Wise Results Status'
print(excel.sheetnames)
sheet.append(['Party name','Leading Votes','Seats Won'])

try:
    source = requests.get("https://results.eci.gov.in/PcResultGenJune2024/index.htm")
    source.raise_for_status()
    soup = BeautifulSoup(source.text, "html.parser")

    parties = soup.find("tbody").find_all("tr", class_="tr")
    # print(parties)
    #print(len(parties))

    for party in parties:
        name = party.find("td", style="text-align:left").text
        # hash_index = name.find("-")
        # sliced_name = name[: hash_index - 1] + " (" + name[hash_index + 2 :] + ")"
        # print(name)

        won = party.find('a').text

        leading = 0
        print(name,leading,won)
        sheet.append([name,leading,won])



except Exception as e:
    print()

excel.save('Parliamentary Constituency General.csv')